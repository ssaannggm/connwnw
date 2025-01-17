#pyinstaller --onefile --noconsole --icon=myicon.ico field_input.py
#region UI / 필드 엑셀내보내기
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side
from pyhwpx import Hwp


#데이터를 저장할 전역 변수
field_list = []
unique_field_list = []
fields_data = []

# tkinter UI 설정
root = tk.Tk()
root.title("필드 추출 / 필드 입력 프로그램 V1.0")

# 버튼 크기를 결정하고 창 크기를 자동으로 맞추기
root.geometry("")  # 빈 값으로 설정하면 자동으로 크기 조정됨
root.grid_propagate(True)  # grid 안의 요소 크기에 맞춰 창 크기 변경

#region 한글파일 업로드 - 필드 추출 - xlsx 저장
def create_excel_file(file_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "Fields Data"

    # 첫 번째 행에 제목 추가
    ws.append(["Fields", "내용"])  # A1 = "Fields", B1 = "내용"

    # 나머지 데이터를 A열에 추가
    for field in unique_field_list:
        ws.append([field])  # 데이터는 A열에 추가됨

    # 1행 배경색 설정 (회색)
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    ws["A1"].fill = gray_fill
    ws["B1"].fill = gray_fill

    # 열 너비 설정 (약 220픽셀에 해당)
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25

    # 열 표시 형식을 텍스트로 설정
    for col in ['A', 'B']:
        for cell in ws[col]:
            cell.number_format = '@'  # 텍스트 형식 지정

    # 테두리 스타일 정의
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # 모든 셀에 테두리 추가
    for row in ws.iter_rows(min_row=1, max_row=len(unique_field_list) + 1, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin_border

    # 파일 저장
    wb.save(file_name)
def save_file():
    file_name = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        title="필드를 추출한 엑셀 파일을 저장하세요"
    )
    if file_name:
        try:
            create_excel_file(file_name)
            messagebox.showinfo("성공", f"파일이 저장되었습니다: {file_name}")
        except Exception as e:
            messagebox.showerror("오류", f"파일 저장 중 오류가 발생했습니다:\n{e}")
    else:
        messagebox.showwarning("취소", "파일 저장이 취소되었습니다.")
def load():
    global field_list
    global unique_field_list
    hwp = Hwp(visible=False)
    try:
        if hwp.Run("FileOpen"):
            field_list = hwp.get_field_list(0).split('\x02') #모든 필드 추출
            unique_field_list = list(set(field_list)) # 중복 삭제
            save_file()
    except Exception as e:
        print(f"파일 선택 중 오류: {e}")
save_button = tk.Button(root, text="[1] 한글파일 업로드 -> 필드 xlsx로 추출", command=load, font=("Arial", 12))
save_button.grid(row=0, column=0, padx=5, pady=5, sticky="ew")
#endregion

#######################################################################################

#region xlsx 업로드 - 한글 파일에 필드 입력
def 필드넣기():
    hwp2 = Hwp()
    try:
        if hwp2.Run("FileOpen"):
            for key, value in fields_data:
                hwp2.PutFieldText(key, str(value) if value is not None else "")
            if hwp2.FileSaveAs():
                messagebox.showinfo("성공", "저장 성공")
    except Exception as e:
        messagebox.showinfo(f"파일 선택 중 오류: {e}")
        print(f"{e}")
def load_file():
    global fields_data
    # 엑셀 파일 열기
    file_name = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        title="내용을 입력한 엑셀 파일을 선택하세요"
    )
    if not file_name:
        messagebox.showwarning("취소", "파일 불러오기가 취소되었습니다.")
        return

    try:
        # 엑셀 파일 읽기
        wb = load_workbook(file_name)
        ws = wb.active

        # 첫 번째 행에서 필드와 내용 열의 위치 찾기
        header = [cell for cell in ws[1]]  # 첫 번째 행의 셀 값 가져옴
        fields_index = None
        content_index = None

        for i, cell in enumerate(header):
            if cell.value == "Fields":
                fields_index = i
            elif cell.value == "내용":
                content_index = i

        if fields_index is None or content_index is None:
            raise ValueError("Fields 또는 내용 열을 찾을 수 없습니다.")

        # 데이터 가져오기
        fields_data = [
            (row[fields_index], row[content_index])
            for row in ws.iter_rows(min_row=2, values_only=True)
        ]

        messagebox.showinfo("성공", f"{file_name}에서 데이터를 성공적으로 불러왔습니다.")
        print(fields_data)  # 디버깅용 출력 (필요시 삭제)
        필드넣기()

    except Exception as e:
        messagebox.showerror("오류", f"파일을 불러오는 중 오류가 발생했습니다:\n{e}")
load_button = tk.Button(root, text="[2] xlsx 업로드 - 한글 파일에 필드 입력", command=load_file, font=("Arial", 12))
load_button.grid(row=1, column=0,padx=5, pady=5, sticky="ew")
#endregion

#######################################################################################

#설명서
def show_manual():
    manual_window = tk.Toplevel(root)
    manual_window.title("사용법")
    manual_window.geometry("1000x600")

    # Text 위젯 추가
    text_widget = tk.Text(manual_window, wrap="word", font=("Arial", 14))
    text_widget.pack(expand=True, fill="both", padx=10, pady=10)

    # 설명서 내용 추가
    manual_content = """
    사용 방법 :
    [1] 한글파일 업로드 -> 필드 xlsx로 추출 버튼 :
        1. (사용자 선택) 셀필드나 필드(력누름틀)를 입력한 한글 파일을 업로드한다. (윈도우 창에서 선택)
        2. (자동)       해당 한글 파일에서 모든 필드를 추출한다.    (중복된 필드는 1개 빼고 전부 삭제)
        3. (사용자 선택) 1행:필드명으로 담긴 xlsx파일로 저장한다.    (다른 이름으로 저장)
    
    [2] xlsx 업로드 - 한글 파일에 필드 입력 버튼 :
        1. (사용자 선택) 내용을 입력한 xlsx파일을 불러온다.
        2. (자동)       xlsx파일 1행(fiels) 2행(내용)을 매칭하여 내부 리스트에 튜플로 저장한다.
        3. (사용자 선택) 필드를 입력한 한글 파일을 선택한다.
        4. (자동)       필드명에 따라 내용을 입한다.
        5. (사용자 선택) 다른 이름으로 저장 기능이 켜지면 저장한다.

    주의사항:
    - 파일 형식 .xlsx
    - 작업 중 문제가 발생하면 Ctrl+Alt+Del로 '한글' 프로그램을 강제로 종료.
    - 문제가 지속되면 프로그램을 재시작.
    - 추출한 xlsx의 필드를 '삭제'가능함.
    - 추출한 xlsx에 필드를 임의로 추가해도 오류는 안남.
    
    by SangMin
    """
    text_widget.insert("1.0", manual_content)

    # 텍스트 위젯을 읽기 전용으로 설정
    text_widget.config(state="disabled")

    # 스크롤바 추가
    scrollbar = tk.Scrollbar(manual_window, orient="vertical", command=text_widget.yview)
    text_widget.config(yscrollcommand=scrollbar.set)
    scrollbar.pack(side="right", fill="y")


# 버튼으로 설명서 창 열기
manual_button = tk.Button(root, text="사용법 보기", command=show_manual, font=("Arial", 12))
manual_button.grid(row=2, column=0,padx=5, pady=5, sticky="ew")


# UI 실행
root.mainloop()
#endregion


