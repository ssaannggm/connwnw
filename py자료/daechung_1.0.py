# %%
#pyinstaller --onefile --noconsole --icon=myicon.ico daechung_1.0.py
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side
from pyhwpx import Hwp
import fitz  # PyMuPDF
from PIL import Image, ImageEnhance
import os
import tempfile  # 임시 파일 생성용
import datetime

#region Tkinter GUI 생성 초기화
root = tk.Tk()
root.title("대충만든 자동 Project_1.0")

# 버튼 크기를 결정하고 창 크기를 자동으로 맞추기
root.geometry("")  # 빈 값으로 설정하면 자동으로 크기 조정됨
root.grid_propagate(True)  # grid 안의 요소 크기에 맞춰 창 크기 변경

var1 = tk.IntVar(value=1)
check1 = tk.Checkbutton(root, text="한글작업 보이기", variable=var1)
check1.grid(row=0, column=0,  padx=5, pady=5,sticky="w")
#endregion

#######################################################################################

#region PDF-이미지추출 초기화
frame1 = tk.Frame(root, highlightbackground="black", highlightthickness=1, padx=10, pady=10)
frame1.grid(row=1, column=0, padx=15, pady=5, sticky="ew")
frame1_label = tk.Label(frame1, text="PDF-이미지 추출", font=("Arial", 12, "bold")) # 기본 정렬 / 전체, 2단 채우기, 입실론, 그림 캡션
frame1_label.grid(row=0, column=0, columnspan=2, padx=5, pady=5,sticky="we")
# 설명 레이블
tk.Label(frame1, text="구역당 페이지 수 ").grid(row=11, column=0, padx=5, pady=5)
entry_pages_per_position = tk.Entry(frame1)
entry_pages_per_position.grid(row=11, column=1, padx=5, pady=5)

tk.Label(frame1, text="출력할 페이지 번호 (예: 2,4)").grid(row=12, column=0, padx=5, pady=5)
entry_target_pages = tk.Entry(frame1)
entry_target_pages.grid(row=12, column=1, padx=5, pady=5)

# 최적화 설정
zoom_x = 4.0  # X축 확대 배율
zoom_y = 4.0  # Y축 확대 배율
mat = fitz.Matrix(zoom_x, zoom_y)  # 확대 매트릭스 설정
#xy축 출력될 곳 비율설정
# DPI 입력
tk.Label(frame1, text="DPI (해상도)").grid(row=13, column=0, padx=5, pady=5)
entry_dpi = tk.Entry(frame1)
entry_dpi.insert(0, "200")  # 기본값
entry_dpi.grid(row=13, column=1, padx=5, pady=5)

# XY 시작/끝 비율 입력
tk.Label(frame1, text="X 시작 비율 (0~1)").grid(row=14, column=0, padx=5, pady=5)
entry_x_start = tk.Entry(frame1)
entry_x_start.insert(0, "0.028")  # 기본값
entry_x_start.grid(row=14, column=1, padx=5, pady=5)

tk.Label(frame1, text="Y 시작 비율 (0~1)").grid(row=15, column=0, padx=5, pady=5)
entry_y_start = tk.Entry(frame1)
entry_y_start.insert(0, "0.032")  # 기본값
entry_y_start.grid(row=15, column=1, padx=5, pady=5)

tk.Label(frame1, text="X 끝 비율 (0~1)").grid(row=16, column=0, padx=5, pady=5)
entry_x_end = tk.Entry(frame1)
entry_x_end.insert(0, "0.972")  # 기본값
entry_x_end.grid(row=16, column=1, padx=5, pady=5)

tk.Label(frame1, text="Y 끝 비율 (0~1)").grid(row=17, column=0, padx=5, pady=5)
entry_y_end = tk.Entry(frame1)
entry_y_end.insert(0, "0.968")  # 기본값
entry_y_end.grid(row=17, column=1, padx=5, pady=5)

tk.Label(frame1, text="타이틀 : ").grid(row=10, column=0, padx=5, pady=5)
entry_title = tk.Entry(frame1)
entry_title.grid(row=10, column=1, padx=5, pady=5)
def show_loading_window(message):
    loading_window = tk.Toplevel(root)
    loading_window.title("작업 진행 중")
    tk.Label(loading_window, text=message, padx=20, pady=20).pack()
    loading_window.update()
    return loading_window
#endregion

#######################################################################################

#region 필드 넣기 초기화

#데이터를 저장할 전역 변수
field_list = []
unique_field_list = []
fields_data = []

frame2 = tk.Frame(root, highlightbackground="black", highlightthickness=1, padx=10, pady=10)
frame2.grid(row=2, column=0, padx=15, pady=5, sticky="ew")
frame2_label = tk.Label(frame2, text="필드 추출, 입력", font=("Arial", 12, "bold")) # 기본 정렬 / 전체, 2단 채우기, 입실론, 그림 캡션
frame2_label.grid(row=0, column=0, columnspan=2, padx=5, pady=5,sticky="we")
#endregion

#######################################################################################

#region 설명서
def show_manual():
    manual_window = tk.Toplevel(root)
    manual_window.title("사용법")
    manual_window.geometry("1000x600")

    # Text 위젯 추가
    text_widget = tk.Text(manual_window, wrap="word", font=("Arial", 14))
    text_widget.pack(expand=True, fill="both", padx=10, pady=10)

    # 설명서 내용 추가
    manual_content = """
    사용 방법 :
    ##### PDF 이미지 템플릿에 넣기 #####
    - 타이틀 : (ex 시스템 동바리)
    - 구역당 페이지 수 : 만약 1-1, 1-2, 1-3, 1-4 가 같은 장소라면 '4'를 입력
    - 출력할 페이지 번호 : 매 구역당 2, 4 쪽을 출력할 거라면 '2,4' 를 입력 (, 기준으로 구별함)
    - '구역당 페이지 수'에 전체 쪽수를 넣으면? : '출력할 페이지 번호' = 쪽수

    PDF 이미지로 저장 버튼: 
        1. PDF 파일 선택
        2. 저장 폴더 선택
    PDF 한글 템플릿 버튼:
        1. PDF 파일 선택
        2. 템플릿 폴더의 미리 작성한 템플릿을 수정해서 프로그램과 같은 폴더에 저장해줌.
        3. 템플릿은 수정가능함.

    ##### 필드 관련 #####
    - 파일 형식 .xlsx
    - 작업 중 문제가 발생하면 Ctrl+Alt+Del로 '한글' 프로그램을 강제로 종료.
    - 추출한 xlsx의 필드를 '삭제'가능함.
    - 추출한 xlsx에 필드를 임의로 추가해도 오류는 안남.
    [1] 한글파일 업로드 -> 필드 xlsx로 추출 버튼 :
        1. (사용자 선택) 필드를 입력한 한글 파일을 업로드한다.
        2. (자동)       해당 한글 파일에서 모든 필드를 추출한다.    (중복된 필드는 1개 빼고 전부 삭제)
        3. (사용자 선택) 1행:필드명으로 담긴 xlsx파일로 저장한다.    (다른 이름으로 저장)
    
    [2] xlsx 업로드 - 한글 파일에 필드 입력 버튼 :
        1. (사용자 선택) 내용을 입력한 xlsx파일을 불러온다.
        2. (자동)       xlsx파일 1행(fiels) 2행(내용)으로 인식 (나머지 무시됨)
        3. (사용자 선택) 필드를 입력한 한글 파일을 선택한다.
        4. (자동)       필드명에 따라 내용을 입한다.
        5. (사용자 선택) 다른 이름으로 저장 기능이 켜지면 저장한다.
    
    by SangMin
    """
    text_widget.insert("1.0", manual_content)

    # 텍스트 위젯을 읽기 전용으로 설정
    text_widget.config(state="disabled")

    # 스크롤바 추가
    scrollbar = tk.Scrollbar(manual_window, orient="vertical", command=text_widget.yview)
    text_widget.config(yscrollcommand=scrollbar.set)
    scrollbar.pack(side="right", fill="y")
# 버튼으로 설명서 창 열기
manual_button = tk.Button(root, text="사용법 보기", command=show_manual, font=("Arial", 11))
manual_button.grid(row=0, column=0,padx=5, pady=5, sticky="e")
#endregion

#######################################################################################

#region PDF 이미지 추출
def extract_images():
    loading_window = None
    try:
        # 사용자 입력 값 가져오기
        pages_per_position = int(entry_pages_per_position.get())
        target_pages = list(map(int, entry_target_pages.get().split(',')))
        pdf_path = filedialog.askopenfilename(title="PDF 파일을 선택하세요", filetypes=[("PDF Files", "*.pdf")])
        if not pdf_path:
            messagebox.showerror("오류", "PDF 파일을 선택하세요.")
            return
        output_folder = filedialog.askdirectory(title="이미지를 저장할 폴더를 선택하세요")
        if not output_folder:
            messagebox.showerror("오류", "이미지를 저장할 폴더를 선택하세요.")
            return
        
        # PDF 열기
        doc = fitz.open(pdf_path)
        pdf_name = os.path.splitext(os.path.basename(pdf_path))[0]
        #with fitz.open(pdf_path) as doc:
        #    pdf_name = os.path.splitext(os.path.basename(pdf_path))[0]

        # 입력값 검증
        dpi = int(entry_dpi.get())
        x_start = float(entry_x_start.get())
        y_start = float(entry_y_start.get())
        x_end = float(entry_x_end.get())
        y_end = float(entry_y_end.get())

        if not (0 <= x_start <= 1 and 0 <= y_start <= 1 and 0 <= x_end <= 1 and 0 <= y_end <= 1):
            raise ValueError("XY 비율 값은 0과 1 사이여야 합니다.")
        if x_start >= x_end or y_start >= y_end:
            raise ValueError("시작 비율은 끝 비율보다 작아야 합니다.")

        loading_window = show_loading_window("PDF 이미지 저장 중입니다. 잠시만 기다려주세요...")
        # 페이지 추출 및 저장
        for i in range(len(doc) // pages_per_position):
            for page_offset in target_pages:
                page_num = (i * pages_per_position) + (page_offset - 1)
                if page_num < len(doc):
                    page = doc[page_num]
                    page_size = page.rect
                    crop_rect = fitz.Rect(
                        page_size.width * x_start, page_size.height * y_start,
                        page_size.width * x_end, page_size.height * y_end
                    )
                    pix = page.get_pixmap(dpi=dpi, colorspace=fitz.csRGB, clip=crop_rect)
                    output_path = os.path.join(output_folder, f"{pdf_name}_page_{page_num + 1}_cropped.png")
                    pix.save(output_path)
            
                    # **이미지 대비 강화 (Pillow 사용)**
                    # 저장된 이미지를 읽고 대비를 강화
                    image = Image.open(output_path)
                    # 대비 및 선명도 강화
                    contrast_enhancer = ImageEnhance.Contrast(image)
                    image = contrast_enhancer.enhance(1.8)  # 대비 강화
                    sharpness_enhancer = ImageEnhance.Sharpness(image)
                    image = sharpness_enhancer.enhance(2.0)  # 선명도 강화
            
                    # 강화된 이미지 저장
                    image.save(output_path)
        doc.close()
        loading_window.destroy()
        messagebox.showinfo("완료", f"이미지 추출 완료!\n저장 폴더: {output_folder}")
        
    except Exception as e:
        messagebox.showerror("오류", f"오류 발생: {str(e)}")
    finally:
        if loading_window:
            loading_window.destroy()
tk.Button(frame1, text="PDF 이미지로 저장", command=extract_images).grid(row=100, column=0,  pady=10)
#endregion

#######################################################################################

#region PDF->한글
def 추출하고템플릿넣기():
    loading_window = None
    temp_files = []  # 임시 파일 경로를 저장할 리스트
    try:
        
        # 사용자 입력 값 가져오기
        pages_per_position = int(entry_pages_per_position.get())
        target_pages = list(map(int, entry_target_pages.get().split(',')))
        pdf_path = filedialog.askopenfilename(title="PDF 파일을 선택하세요", filetypes=[("PDF Files", "*.pdf")])
        if not pdf_path:
            messagebox.showerror("오류", "PDF 파일을 선택하세요.")
            return
        
        loading_window = show_loading_window("PDF 추출 및 한글 템플릿 작성 중입니다. 잠시만 기다려주세요...")

        # PDF 열기
        doc = fitz.open(pdf_path)
                 
        # 입력값 검증
        dpi = int(entry_dpi.get())
        x_start = float(entry_x_start.get())
        y_start = float(entry_y_start.get())
        x_end = float(entry_x_end.get())
        y_end = float(entry_y_end.get())
        타이틀 = str(entry_title.get())
        if not (0 <= x_start <= 1 and 0 <= y_start <= 1 and 0 <= x_end <= 1 and 0 <= y_end <= 1):
            raise ValueError("XY 비율 값은 0과 1 사이여야 합니다.")
        if x_start >= x_end or y_start >= y_end:
            raise ValueError("시작 비율은 끝 비율보다 작아야 합니다.")

        # 현재 날짜와 시간 가져오기
        now = datetime.datetime.now()
        formatted_date = now.strftime("%y.%m.%d.")  # 날짜: YY.MM.DD.
        

        # 파일 이름 생성
        file_name = f"{타이틀}_({formatted_date}).hwp"

        # 출력될 총 페이지 수 계산
        total_pdf_pages = len(doc)
        total_sections = total_pdf_pages // pages_per_position
        total_output_pages = total_sections * len(target_pages)
        복사할표갯수 = total_output_pages//2 - 1
        if var1.get() == 1: hwp = Hwp(visible=True)
        else: hwp = Hwp(visible=False)

        hwp.Open(r"템플릿\설치위치_도면_템플릿.hwp")

        #템플릿 - 도면 표 복사
        hwp.get_into_nth_table(-1) #마지막 표로 진입
        hwp.SelectCtrlFront() #진입한 표 선택
        hwp.Copy() #복
        for i in range(복사할표갯수):
            hwp.Paste()
        
        #템플릿 - 목차 표 복사 
        hwp.get_into_nth_table(0)#첫번째 표
        hwp.SelectCtrlFront() #진입한 표 선택
        hwp.get_into_nth_table(0)#첫번째 표
        hwp.SelectCtrlFront() #진입한 표 선택
        hwp.TableCellBlock()
        hwp.TableColPageDown()
        hwp.TableCellBlockExtend()
        hwp.TableColEnd()
        hwp.Copy() #복
        def 타이틀_표_늘리기():
            pset = hwp.HParameterSet.HSelectionOpt
            hwp.HAction.GetDefault("Paste", pset.HSet)
            pset.option = 3
            hwp.HAction.Execute("Paste", pset.HSet)
            hwp.HAction.Run("TableLowerCell")
        def 윗선얇게():
            pset = hwp.HParameterSet.HCellBorderFill
            hwp.HAction.Run("TableCellBlockExtend")
            hwp.HAction.Run("TableColBegin")
            hwp.HAction.GetDefault("CellBorder", pset.HSet)
            pset.BorderWidthTop = hwp.HwpLineWidth("0.12mm")
            pset.BorderTypeTop = hwp.HwpLineType("Solid")
            hwp.HAction.Execute("CellBorder", pset.HSet)
            hwp.HAction.Run("TableColEnd")
            hwp.HAction.Run("Cancel")
            hwp.HAction.Run("TableCellBlock")
        for i in range(복사할표갯수):
            타이틀_표_늘리기()
            윗선얇게()
        
        hwp.PutFieldText("타이틀", 타이틀)#타이틀 입력

        # 이미지 추출 및 저장
        image_index = 0  # 이미지를 고유하게 식별할 인덱스 변수 초기화
        
        # 페이지 추출 및 저장
        for i in range(len(doc) // pages_per_position): # 구역별
            hwp.PutFieldText("표번호1", f"{(i+1)}", i)
            hwp.PutFieldText("표번호2", f"{(i+1)}", i)
            hwp.PutFieldText("번호", f"{(i+1)}", i)
            for page_offset in target_pages: #페이지 별
                page_num = (i * pages_per_position) + (page_offset - 1)
                if page_num < len(doc):
                    page = doc[page_num]
                    page_size = page.rect
                    
                    # 특정 영역의 텍스트 추출
                    clip_rect = fitz.Rect(
                        page_size.width * 0.28, 
                        page_size.height * 0.9,
                        page_size.width * 0.72, 
                        page_size.height * 1
                    )
                    location_text = page.get_text("text", clip=clip_rect)
                    hwp.PutFieldText("설치위치", location_text, i)
                    hwp.PutFieldText("표설치위치1", location_text, i)
                    hwp.PutFieldText("표설치위치2", location_text, i)

                    #영역 잘라내기
                    crop_rect = fitz.Rect(
                        page_size.width * x_start, page_size.height * y_start,
                        page_size.width * x_end, page_size.height * y_end
                    )
                    pix = page.get_pixmap(dpi=dpi, colorspace=fitz.csRGB, clip=crop_rect)

                    # 임시 파일 생성 및 저장
                    temp_file = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
                    temp_path = temp_file.name
                    temp_file.close()
                    temp_files.append(temp_path)  # 임시 파일 경로 저장
                    pix.save(temp_path)

                    # 이미지 대비 및 선명도 강화
                    with Image.open(temp_path) as image:
                        contrast_enhancer = ImageEnhance.Contrast(image)
                        enhanced_image = contrast_enhancer.enhance(1.8)  # 대비 강화
                        sharpness_enhancer = ImageEnhance.Sharpness(enhanced_image)
                        enhanced_image = sharpness_enhancer.enhance(2.0)  # 선명도 강화
                        enhanced_image.save(temp_path)

                    # 사진 입력
                    try:
                        hwp.move_to_field("도면",image_index)
                        hwp.insert_picture(temp_path, sizeoption=3)
                    except Exception as e:
                        print(f"API 호출 중 오류 발생: {e}")
                    
                    #
                    # 이미지 인덱스 증가
                    
                    image_index += 1
        
        hwp.SaveAs(file_name)
        doc.close()
        loading_window.destroy()
        messagebox.showinfo("완료", f"이미지 추출 완료!\n저장:{file_name}")
    except Exception as e:
        messagebox.showerror("오류", f"오류 발생: {str(e)}")
    finally:
        if loading_window:
            loading_window.destroy()
        hwp.clear()
        hwp.quit()
        # 작업 완료 후 모든 임시 파일 삭제
        for temp_path in temp_files:
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except Exception as e:
                    print(f"임시 파일 삭제 오류: {e}")
tk.Button(frame1, text="PDF->한글 템플릿", command=추출하고템플릿넣기).grid(row=100, column=1, pady=10)
#endregion

#######################################################################################

#region 한글파일 업로드 - 필드 추출 - xlsx 저장
def create_excel_file(file_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "Fields Data"

    # 첫 번째 행에 제목 추가
    ws.append(["Fields", "내용"])  # A1 = "Fields", B1 = "내용"

    # 나머지 데이터를 A열에 추가
    for field in unique_field_list:
        ws.append([field])  # 데이터는 A열에 추가됨

    # 1행 배경색 설정 (회색)
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    ws["A1"].fill = gray_fill
    ws["B1"].fill = gray_fill

    # 열 너비 설정 (약 220픽셀에 해당)
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25

    # 열 표시 형식을 텍스트로 설정
    for col in ['A', 'B']:
        for cell in ws[col]:
            cell.number_format = '@'  # 텍스트 형식 지정

    # 테두리 스타일 정의
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # 모든 셀에 테두리 추가
    for row in ws.iter_rows(min_row=1, max_row=len(unique_field_list) + 1, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin_border

    # 파일 저장
    wb.save(file_name)
def save_file():
    file_name = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        title="필드를 추출한 엑셀 파일을 저장하세요"
    )
    if file_name:
        try:
            create_excel_file(file_name)
            messagebox.showinfo("성공", f"파일이 저장되었습니다: {file_name}")
        except Exception as e:
            messagebox.showerror("오류", f"파일 저장 중 오류가 발생했습니다:\n{e}")
    else:
        messagebox.showwarning("취소", "파일 저장이 취소되었습니다.")
def load():
    global field_list
    global unique_field_list
    if var1.get() == 1: hwp = Hwp(visible=True)
    else: hwp = Hwp(visible=False)
    try:
        if hwp.Run("FileOpen"):
            field_list = hwp.get_field_list(0).split('\x02') #모든 필드 추출
            unique_field_list = list(set(field_list)) # 중복 삭제
            save_file()
    except Exception as e:
        print(f"파일 선택 중 오류: {e}")
    finally:
        hwp.clear()
        hwp.quit()
save_button = tk.Button(frame2, text="[1] 한글파일 업로드 -> 필드 xlsx로 추출", command=load, font=("Arial", 11))
save_button.grid(row=11, column=0, padx=5, pady=5, sticky="ew")
#endregion

#######################################################################################

#region xlsx 업로드 - 한글 파일에 필드 입력
def 필드넣기():
    if var1.get() == 1: hwp = Hwp(visible=True)
    else: hwp = Hwp(visible=False)
    try:
        if hwp.Run("FileOpen"):
            for key, value in fields_data:
                hwp.PutFieldText(key, str(value) if value is not None else "")
            if hwp.FileSaveAs():
                messagebox.showinfo("성공", "저장 성공")
    except Exception as e:
        messagebox.showinfo(f"파일 선택 중 오류: {e}")
        print(f"{e}")
    finally:
        hwp.clear()
        hwp.quit()
def load_file():
    global fields_data
    # 엑셀 파일 열기
    file_name = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        title="내용을 입력한 엑셀 파일을 선택하세요"
    )
    if not file_name:
        messagebox.showwarning("취소", "파일 불러오기가 취소되었습니다.")
        return

    try:
        # 엑셀 파일 읽기
        wb = load_workbook(file_name)
        ws = wb.active

        # 첫 번째 행에서 필드와 내용 열의 위치 찾기
        header = [cell for cell in ws[1]]  # 첫 번째 행의 셀 값 가져옴
        fields_index = None
        content_index = None

        for i, cell in enumerate(header):
            if cell.value == "Fields":
                fields_index = i
            elif cell.value == "내용":
                content_index = i

        if fields_index is None or content_index is None:
            raise ValueError("Fields 또는 내용 열을 찾을 수 없습니다.")

        # 데이터 가져오기
        fields_data = [
            (row[fields_index], row[content_index])
            for row in ws.iter_rows(min_row=2, values_only=True)
        ]

        messagebox.showinfo("성공", f"{file_name}에서 데이터를 성공적으로 불러왔습니다.")
        print(fields_data)  # 디버깅용 출력 (필요시 삭제)
        필드넣기()

    except Exception as e:
        messagebox.showerror("오류", f"파일을 불러오는 중 오류가 발생했습니다:\n{e}")
load_button = tk.Button(frame2, text="[2] xlsx 업로드 - 한글 파일에 필드 입력", command=load_file, font=("Arial", 11))
load_button.grid(row=12, column=0,padx=5, pady=5, sticky="ew")
#endregion

#######################################################################################

# Tkinter GUI 실행
root.mainloop()



